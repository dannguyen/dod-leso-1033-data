#!/usr/bin/env python
from sys import path as syspath; syspath.append('./scripts')
from utils import mylog

import csv
import re
from openpyxl import load_workbook
from pathlib import Path
from xlrd import open_workbook

SRC_DIR = Path('data/collected/disp/agencies')



def gather_bookpaths():
    return sorted(SRC_DIR.rglob('*.xlsx'))

def get_csvdir(bookpath):
    """
    :bookpath <str> or <Path>
        something like "data/collected/disp/agencies/2020-03-31/all-states.xlsx"

    Returns: <Path>
        data/collected/disp/agencies/2020-03-31/all-states/
    """
    bookpath = Path(bookpath)
    bookdir = bookpath.parent
    csvdir = bookdir.joinpath(bookpath.stem)

    return csvdir





def extract_sheet(sheet):
    """
    :sheet <openpyxl.worksheet.worksheet.Worksheet>

    Returns <list>: a 2D array of values from the sheet
    """
    data = [[cell.value for cell in row] for row in sheet.rows]
    return data


def extract_all_sheets(srcpath):
    """
    :srcpath <str> or <Path>: path to a spreadsheet file

    Returns: <dict> each key is a sheet name, each value is the corresponding 2D-array of values
    """
    srcpath = Path(srcpath)
    mylog(f"Loading workbook {srcpath}")

    # if re.search(r'\.xls$', str(srcpath)):
    #     # old format; use xlrd
    #     # note that integers are wrongly typecasted to floats, but oh well
    #     sheet = open_workbook(srcpath).sheets()[0]
    #     ncols = sheet.row_len(0)
    #     nrows = sheet.nrows
    #     return [[sheet.cell_value(i, j) for j in range(ncols)] for i in range(nrows)]
    # else:

    book = load_workbook(srcpath, read_only=True)
    mylog(f"...loaded: {len(book.sheetnames)} sheets")

    data = {}

    for sname in book.sheetnames:
        data[sname] = extract_sheet(book[sname])

    return data


def main():
    bookpaths = gather_bookpaths()
    mylog(f"Found {len(bookpaths)}...")
    for srcpath in bookpaths:

        destdir = get_csvdir(srcpath)
        # test if files already exists
        childs = list(destdir.glob('*.csv'))
        if destdir.is_dir() and len(childs) > 0:
            mylog(f"...Skipping {destdir}; {len(childs)} csv files found")
        else:
            destdir.mkdir(exist_ok=True, parents=True)
            data = extract_all_sheets(srcpath)


            for sheetname, rows in data.items():
                destpath = destdir.joinpath(f'{sheetname}.csv')

                with open(destpath, 'w') as dest:
                    outs = csv.writer(dest)
                    outs.writerows(rows)
                    mylog(f"...wrote {len(rows)} rows to: {destpath}")

# def testfoo():
#     data = extract_all_sheets(X_PATH)
#     destdir = Path('/tmp/testcsvfoo')
#     destdir.mkdir(exist_ok=True, parents=True)
#     for sheetname, rows in data.items():
#         destpath = destdir.joinpath(f'{sheetname}.csv')
#         with open(destpath, 'w') as dest:
#             outs = csv.writer(dest)
#             outs.writerows(rows)
#             mylog(f"...wrote {len(rows)} rows to: {destpath}")


if __name__ == '__main__':
    main()
